from decimal import Decimal
from datetime import timedelta
import csv
import io
import json
import secrets
from urllib.parse import urlencode

from django.contrib.auth import login, logout
from django.contrib.auth.hashers import check_password, make_password
from django.contrib.auth.models import User
from django.contrib import messages
from django.conf import settings
from django.core.mail import send_mail
from django.http import Http404, HttpResponse, JsonResponse
from django.db import transaction
from django.db.models import Count, F, Sum, Q
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl import Workbook
import requests

from .forms import CRUD_FORMS, BusinessProfileForm, CashCloseForm, EmailVerificationForm, ExcelImportForm, ExpenseForm, IntegrationConnectionForm, LoginForm, OnboardingForm, PasswordResetRequestForm, PasswordResetVerifyForm, PurchaseForm, RegisterForm, SaleForm, StockMovementForm, StockTransferForm, TeamMemberForm
from .models import (
    AccountProfile,
    ActivityLog,
    BackupLog,
    Branch,
    BusinessProfile,
    Category,
    Customer,
    EmailVerificationCode,
    Employee,
    Expense,
    HospitalityUnit,
    IntegrationConnection,
    Payment,
    PasswordResetCode,
    Permission,
    Product,
    Purchase,
    PurchaseItem,
    Notification,
    CashRegisterSession,
    Role,
    RolePermission,
    Sale,
    SaleItem,
    StockMovement,
    StockTransfer,
    Supplier,
    UserBusinessAccess,
)


CRUD_CONFIG = {
    'branches': {'model': Branch, 'title': 'Branches', 'icon': 'fa-code-branch', 'fields': ['name', 'location', 'phone', 'is_active']},
    'categories': {'model': Category, 'title': 'Categories', 'icon': 'fa-tags', 'fields': ['name']},
    'suppliers': {'model': Supplier, 'title': 'Suppliers', 'icon': 'fa-truck-field', 'fields': ['name', 'phone', 'email', 'location']},
    'products': {'model': Product, 'title': 'Products & Services', 'icon': 'fa-box', 'fields': ['name', 'item_type', 'sku', 'branch', 'category', 'supplier', 'stock', 'price']},
    'customers': {'model': Customer, 'title': 'Customers', 'icon': 'fa-users', 'fields': ['name', 'phone', 'email', 'debt_balance', 'loyalty_points']},
    'employees': {'model': Employee, 'title': 'Employees', 'icon': 'fa-id-badge', 'fields': ['name', 'role', 'branch', 'phone', 'salary']},
    'hospitality': {'model': HospitalityUnit, 'title': 'Rooms & Tables', 'icon': 'fa-hotel', 'fields': ['name', 'unit_type', 'status', 'branch', 'current_guest']},
    'payments': {'model': Payment, 'title': 'Payments', 'icon': 'fa-mobile-screen', 'fields': ['method', 'branch', 'amount', 'status', 'transaction_reference']},
    'expenses': {'model': Expense, 'title': 'Expenses', 'icon': 'fa-money-bill-trend-up', 'fields': ['title', 'branch', 'category', 'amount', 'spent_at']},
    'transfers': {'model': StockTransfer, 'title': 'Stock Transfers', 'icon': 'fa-right-left', 'fields': ['product_name', 'from_branch', 'to_branch', 'quantity', 'created_at']},
}

MODULES_BY_BUSINESS = {
    'shop': ['pos', 'inventory', 'customers', 'payments', 'reports', 'integrations'],
    'boutique': ['pos', 'inventory', 'customers', 'payments', 'reports', 'integrations'],
    'restaurant': ['pos', 'inventory', 'customers', 'payments', 'employees', 'hospitality', 'expenses', 'reports', 'integrations'],
    'pharmacy': ['pos', 'inventory', 'customers', 'payments', 'suppliers', 'reports', 'integrations'],
    'hardware': ['pos', 'inventory', 'customers', 'payments', 'suppliers', 'transfers', 'reports', 'integrations'],
    'hotel': ['pos', 'inventory', 'customers', 'payments', 'employees', 'hospitality', 'expenses', 'reports', 'integrations'],
    'mini_market': ['pos', 'inventory', 'customers', 'payments', 'employees', 'expenses', 'reports', 'integrations'],
    'wholesale': ['pos', 'inventory', 'customers', 'payments', 'suppliers', 'transfers', 'reports', 'integrations'],
    'salon': ['pos', 'customers', 'payments', 'employees', 'expenses', 'reports', 'integrations'],
    'clinic': ['customers', 'payments', 'employees', 'expenses', 'reports', 'integrations'],
    'garage': ['pos', 'inventory', 'customers', 'payments', 'employees', 'expenses', 'suppliers', 'reports', 'integrations'],
    'laundry': ['pos', 'customers', 'payments', 'employees', 'expenses', 'reports', 'integrations'],
    'construction': ['customers', 'payments', 'employees', 'expenses', 'suppliers', 'reports', 'integrations'],
    'agency': ['customers', 'payments', 'employees', 'expenses', 'reports', 'integrations'],
    'school': ['customers', 'employees', 'payments', 'expenses', 'reports', 'integrations'],
    'club': ['customers', 'employees', 'payments', 'expenses', 'reports', 'integrations'],
    'other': ['pos', 'inventory', 'customers', 'payments', 'employees', 'expenses', 'hospitality', 'suppliers', 'transfers', 'reports', 'integrations'],
}

DEFAULT_PERMISSIONS = [
    ('manage_business', 'Manage business setup'),
    ('manage_team', 'Manage team and roles'),
    ('manage_sales', 'Record sales and services'),
    ('manage_inventory', 'Manage products, services, and stock'),
    ('manage_customers', 'Manage customers'),
    ('manage_payments', 'Manage payments'),
    ('view_reports', 'View reports'),
]

DEFAULT_ROLES = {
    'owner': {
        'display_name': 'Owner',
        'permissions': [code for code, _ in DEFAULT_PERMISSIONS],
    },
    'manager': {
        'display_name': 'Manager',
        'permissions': ['manage_sales', 'manage_inventory', 'manage_customers', 'manage_payments', 'view_reports'],
    },
    'cashier': {
        'display_name': 'Cashier',
        'permissions': ['manage_sales', 'manage_customers', 'manage_payments'],
    },
    'stock_controller': {
        'display_name': 'Stock Controller',
        'permissions': ['manage_inventory', 'view_reports'],
    },
    'accountant': {
        'display_name': 'Accountant',
        'permissions': ['manage_payments', 'view_reports'],
    },
}


def money(value):
    return value or Decimal('0.00')


def ensure_default_roles_and_permissions():
    permission_map = {}
    for code, name in DEFAULT_PERMISSIONS:
        permission_map[code], _ = Permission.objects.get_or_create(code=code, defaults={'name': name})

    for role_name, config in DEFAULT_ROLES.items():
        role, _ = Role.objects.get_or_create(name=role_name, defaults={'display_name': config['display_name']})
        for permission_code in config['permissions']:
            RolePermission.objects.get_or_create(role=role, permission=permission_map[permission_code])


def link_pending_accesses(user):
    if not user.email:
        return
    UserBusinessAccess.objects.filter(email__iexact=user.email, user__isnull=True).update(user=user)


def account_is_verified(user):
    profile, _ = AccountProfile.objects.get_or_create(user=user)
    return profile.email_verified

def send_email_verification_code(user):
    AccountProfile.objects.get_or_create(user=user)
    EmailVerificationCode.objects.filter(user=user, used_at__isnull=True).update(used_at=timezone.now())
    code = f'{secrets.randbelow(1000000):06d}'
    EmailVerificationCode.objects.create(
        user=user,
        email=user.email,
        code_hash=make_password(code),
        expires_at=timezone.now() + timedelta(minutes=15),
    )
    send_mail(
        subject='Verify your SMS Suite account',
        message=f'Your SMS Suite verification code is {code}. It expires in 15 minutes.',
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
        fail_silently=settings.DEBUG,
    )
    return code if settings.DEBUG else ''


def user_can_access_business(user, business):
    if not user.is_authenticated:
        return False
    if business.owner_email and user.email and business.owner_email.lower() == user.email.lower():
        return True
    return UserBusinessAccess.objects.filter(user=user, business=business, is_active=True).exists()


def accessible_businesses(user):
    if not user.is_authenticated:
        return BusinessProfile.objects.none()
    business_ids = UserBusinessAccess.objects.filter(user=user, is_active=True).values_list('business_id', flat=True)
    queryset = BusinessProfile.objects.filter(pk__in=business_ids)
    if user.email:
        queryset = queryset | BusinessProfile.objects.filter(owner_email__iexact=user.email)
    return queryset.distinct().order_by('name')


def user_has_business_permission(request, permission_code):
    profile = get_profile_for_request(request)
    if profile.owner_email and request.user.email and profile.owner_email.lower() == request.user.email.lower():
        return True
    return UserBusinessAccess.objects.filter(
        user=request.user,
        business=profile,
        is_active=True,
        role__role_permissions__permission__code=permission_code,
    ).exists()


def permission_denied(permission_code):
    return HttpResponse(f'You do not have permission for this action: {permission_code}', status=403)


def resource_permission(resource):
    return {
        'branches': 'manage_business',
        'categories': 'manage_inventory',
        'suppliers': 'manage_inventory',
        'products': 'manage_inventory',
        'customers': 'manage_customers',
        'employees': 'manage_team',
        'hospitality': 'manage_business',
        'payments': 'manage_payments',
        'expenses': 'manage_payments',
        'transfers': 'manage_inventory',
    }.get(resource, 'manage_business')


def dashboard(request):
    profile = get_profile_for_request(request)
    modules = profile.selected_modules or MODULES_BY_BUSINESS.get(profile.business_type, MODULES_BY_BUSINESS['shop'])
    sales_total = money(SaleItem.objects.filter(sale__branch__profile=profile).aggregate(total=Sum(F('quantity') * F('unit_price')))['total'])
    profit_total = money(SaleItem.objects.filter(sale__branch__profile=profile).aggregate(total=Sum(F('quantity') * (F('unit_price') - F('unit_cost'))))['total'])
    context = {
        'sales_total': sales_total,
        'profit_total': profit_total,
        'stock_count': Product.objects.filter(branch__profile=profile).aggregate(total=Sum('stock'))['total'] or 0,
        'customer_count': Customer.objects.filter(profile=profile).count(),
        'debt_total': money(Customer.objects.filter(profile=profile).aggregate(total=Sum('debt_balance'))['total']),
        'payments': Payment.objects.filter(branch__profile=profile).select_related('branch')[:5],
        'activity_logs': ActivityLog.objects.filter(profile=profile).select_related('branch')[:6],
        'notifications': Notification.objects.filter(profile=profile, is_read=False).select_related('branch')[:6],
        'expense_total': money(Expense.objects.filter(branch__profile=profile).aggregate(total=Sum('amount'))['total']),
        'low_stock': Product.objects.filter(branch__profile=profile, stock__lte=F('reorder_level')).select_related('category', 'branch')[:6],
        'recent_sales': Sale.objects.filter(branch__profile=profile).select_related('branch', 'customer', 'cashier').prefetch_related('items')[:5],
        'branch_summary': Branch.objects.filter(profile=profile).annotate(
            product_count=Count('products', distinct=True),
            employee_count=Count('employees', distinct=True),
            sale_count=Count('sales', distinct=True),
        ),
        'hospitality': HospitalityUnit.objects.filter(branch__profile=profile).select_related('branch')[:8],
        'profile': profile,
        'modules': modules,
        'business_label': profile.get_business_type_display(),
    }
    return render(request, 'core/dashboard.html', context)


def get_profile():
    return BusinessProfile.objects.first() or BusinessProfile.objects.create(selected_modules=MODULES_BY_BUSINESS['shop'])


def get_profile_for_request(request):
    profile_id = request.session.get('active_profile_id')
    if profile_id:
        profile = BusinessProfile.objects.filter(pk=profile_id).first()
        if profile and user_can_access_business(request.user, profile):
            return profile
    raise Http404('No active business selected.')


def auth_view(request):
    if request.user.is_authenticated:
        return redirect('workspace')
    login_form = LoginForm(request, data=request.POST or None)
    register_form = RegisterForm()
    if request.method == 'POST' and request.POST.get('mode') == 'login':
        if login_form.is_valid():
            login(request, login_form.get_user())
            link_pending_accesses(login_form.get_user())
            if not account_is_verified(login_form.get_user()):
                debug_code = send_email_verification_code(login_form.get_user())
                if debug_code:
                    request.session['email_verification_debug_code'] = debug_code
                return redirect('verify_email')
            return redirect('workspace')
    return render(request, 'core/auth.html', {'login_form': login_form, 'register_form': register_form})


def register_view(request):
    if request.user.is_authenticated:
        return redirect('workspace')
    register_form = RegisterForm(request.POST or None)
    login_form = LoginForm()
    if request.method == 'POST' and register_form.is_valid():
        user = register_form.save()
        AccountProfile.objects.get_or_create(user=user)
        link_pending_accesses(user)
        debug_code = send_email_verification_code(user)
        login(request, user)
        if debug_code:
            request.session['email_verification_debug_code'] = debug_code
        messages.success(request, 'Account created. Verify your email to protect your business workspace.')
        return redirect('verify_email')
    return render(request, 'core/auth.html', {'login_form': login_form, 'register_form': register_form, 'show_register': True})



def verify_email(request):
    if not request.user.is_authenticated:
        return redirect('auth')
    profile, _ = AccountProfile.objects.get_or_create(user=request.user)
    if profile.email_verified:
        return redirect('workspace')

    debug_code = request.session.get('email_verification_debug_code', '') if settings.DEBUG else ''
    if not EmailVerificationCode.objects.filter(user=request.user, used_at__isnull=True, expires_at__gte=timezone.now()).exists():
        debug_code = send_email_verification_code(request.user)
        if debug_code:
            request.session['email_verification_debug_code'] = debug_code

    if request.method == 'POST' and request.POST.get('resend') == '1':
        debug_code = send_email_verification_code(request.user)
        if debug_code:
            request.session['email_verification_debug_code'] = debug_code
        messages.success(request, 'A new verification code was sent.')
        return redirect('verify_email')

    form = EmailVerificationForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        code = form.cleaned_data['code'].strip()
        verification = EmailVerificationCode.objects.filter(user=request.user, used_at__isnull=True).first()
        if not verification or not verification.is_active:
            form.add_error('code', 'This code is invalid or expired.')
        elif not check_password(code, verification.code_hash):
            verification.attempts = F('attempts') + 1
            verification.save(update_fields=['attempts'])
            form.add_error('code', 'This code is invalid or expired.')
        else:
            verification.used_at = timezone.now()
            verification.save(update_fields=['used_at'])
            profile.email_verified = True
            profile.save(update_fields=['email_verified'])
            request.session.pop('email_verification_debug_code', None)
            messages.success(request, 'Email verified. Your workspace is ready.')
            return redirect('workspace')
    return render(request, 'core/verify_email.html', {'form': form, 'debug_code': debug_code})

def password_reset_request(request):
    if request.user.is_authenticated:
        return redirect('workspace')

    form = PasswordResetRequestForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        email = form.cleaned_data['email'].strip().lower()
        user = User.objects.filter(email__iexact=email, is_active=True).first()
        debug_code = ''
        if user:
            PasswordResetCode.objects.filter(user=user, used_at__isnull=True).update(used_at=timezone.now())
            code = f'{secrets.randbelow(1000000):06d}'
            PasswordResetCode.objects.create(
                user=user,
                email=email,
                code_hash=make_password(code),
                expires_at=timezone.now() + timedelta(minutes=10),
            )
            send_password_reset_code(email, code)
            if settings.DEBUG:
                debug_code = code
        request.session['password_reset_email'] = email
        if debug_code:
            request.session['password_reset_debug_code'] = debug_code
        messages.success(request, 'If that account exists, a verification code has been sent.')
        return redirect('password_reset_verify')

    return render(request, 'core/password_reset.html', {'form': form})


def password_reset_verify(request):
    if request.user.is_authenticated:
        return redirect('workspace')

    initial_email = request.session.get('password_reset_email', '')
    form = PasswordResetVerifyForm(request.POST or None, initial={'email': initial_email})
    debug_code = request.session.get('password_reset_debug_code', '') if settings.DEBUG else ''

    if request.method == 'POST' and form.is_valid():
        email = form.cleaned_data['email'].strip().lower()
        code = form.cleaned_data['code'].strip()
        reset_code = PasswordResetCode.objects.filter(email__iexact=email, used_at__isnull=True).select_related('user').first()
        if not reset_code or not reset_code.is_active:
            form.add_error('code', 'This code is invalid or has expired.')
        elif not check_password(code, reset_code.code_hash):
            reset_code.attempts = F('attempts') + 1
            reset_code.save(update_fields=['attempts'])
            form.add_error('code', 'This code is invalid or has expired.')
        else:
            reset_code.user.set_password(form.cleaned_data['new_password1'])
            reset_code.user.save(update_fields=['password'])
            reset_code.used_at = timezone.now()
            reset_code.save(update_fields=['used_at'])
            request.session.pop('password_reset_email', None)
            request.session.pop('password_reset_debug_code', None)
            messages.success(request, 'Password updated. You can sign in now.')
            return redirect('password_reset_complete')

    return render(request, 'core/password_reset_confirm.html', {'form': form, 'debug_code': debug_code})


def password_reset_complete(request):
    return render(request, 'core/password_reset_complete.html')


def send_password_reset_code(email, code):
    send_mail(
        subject='Your SMS Suite verification code',
        message=(
            'Use this verification code to reset your SMS Suite password:\n\n'
            f'{code}\n\n'
            'This code expires in 10 minutes. If you did not request it, ignore this email.'
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[email],
        fail_silently=False,
    )



def health_check(request):
    return JsonResponse({
        "status": "ok",
        "app": "SMS Suite",
        "time": timezone.now().isoformat(),
    })

def tour(request):
    return render(request, 'core/tour.html')


def logout_view(request):
    logout(request)
    return redirect('auth')


def workspace(request):
    profiles = accessible_businesses(request.user)
    return render(request, 'core/workspace.html', {'profiles': profiles})


def start_new_business(request):
    ensure_default_roles_and_permissions()
    email = request.user.email or ''
    name = request.user.get_full_name() or request.user.username or 'Owner'
    profile = BusinessProfile.objects.create(
        name=f"{name}'s Business",
        owner_email=email,
        selected_modules=MODULES_BY_BUSINESS['shop'],
        setup_completed=False,
    )
    owner_role = Role.objects.get(name='owner')
    UserBusinessAccess.objects.update_or_create(
        user=request.user,
        business=profile,
        defaults={'email': email, 'role': owner_role, 'is_active': True},
    )
    request.session['active_profile_id'] = profile.id
    return redirect('onboarding')


def open_business(request, pk):
    profile = get_object_or_404(accessible_businesses(request.user), pk=pk)
    request.session['active_profile_id'] = profile.id
    return redirect('dashboard' if profile.setup_completed else 'onboarding')


def onboarding(request):
    profile = get_profile_for_request(request)
    if request.method == 'POST':
        form = OnboardingForm(request.POST, instance=profile)
        if form.is_valid():
            profile = form.save(commit=False)
            if not profile.selected_modules:
                profile.selected_modules = MODULES_BY_BUSINESS.get(profile.business_type, MODULES_BY_BUSINESS['shop'])
            profile.setup_completed = True
            profile.save()
            if not profile.trial_started_at:
                profile.trial_started_at = timezone.now()
                profile.trial_ends_at = timezone.now() + timedelta(days=180)
                profile.subscription_status = 'trial'
                profile.save(update_fields=['trial_started_at', 'trial_ends_at', 'subscription_status'])
            Branch.objects.get_or_create(
                profile=profile,
                name=form.cleaned_data.get('branch_name') or 'Main Branch',
                defaults={
                    'location': form.cleaned_data.get('branch_location') or 'Kigali',
                    'phone': form.cleaned_data.get('branch_phone') or profile.momo_number or profile.whatsapp_number,
                },
            )
            Category.objects.get_or_create(profile=profile, name='General')
            ensure_default_connections(profile)
            ensure_default_roles_and_permissions()
            ActivityLog.objects.create(profile=profile, action='Completed first-time business setup')
            messages.success(request, 'Setup completed. Your dashboard is ready.')
            return redirect('dashboard')
    else:
        if not profile.selected_modules:
            profile.selected_modules = MODULES_BY_BUSINESS.get(profile.business_type, MODULES_BY_BUSINESS['shop'])
        form = OnboardingForm(instance=profile)
    return render(request, 'core/onboarding.html', {'form': form, 'profile': profile, 'modules_by_business': MODULES_BY_BUSINESS})


@transaction.atomic
def pos(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_sales'):
        return permission_denied('manage_sales')
    if request.method == 'POST':
        form = SaleForm(request.POST, profile=profile)
        if form.is_valid():
            raw_cart = form.cleaned_data.get('cart_items') or ''
            if raw_cart:
                cart_rows = json.loads(raw_cart)
            else:
                cart_rows = [{'product_id': form.cleaned_data['product'].pk, 'quantity': form.cleaned_data['quantity']}]

            product_ids = [int(row['product_id']) for row in cart_rows]
            products = Product.objects.select_for_update().filter(pk__in=product_ids, branch__profile=profile)
            products_by_id = {product.id: product for product in products}
            normalized_items = []
            subtotal = Decimal('0.00')
            total_profit_points = 0

            for row in cart_rows:
                product_id = int(row['product_id'])
                quantity = int(row['quantity'])
                product = products_by_id.get(product_id)
                if not product:
                    form.add_error(None, 'One product in the cart could not be found.')
                    return render(request, 'core/pos.html', build_pos_context(profile, form))
                if product.track_inventory and quantity > product.stock:
                    form.add_error(None, f'Only {product.stock} units of {product.name} are available.')
                    return render(request, 'core/pos.html', build_pos_context(profile, form))
                line_total = product.price * quantity
                subtotal += line_total
                total_profit_points += int(line_total // 10)
                normalized_items.append((product, quantity))

            discount = form.cleaned_data.get('discount_amount') or Decimal('0.00')
            amount_paid = form.cleaned_data.get('amount_paid') or Decimal('0.00')
            grand_total = max(subtotal - discount, Decimal('0.00'))
            balance_due = max(grand_total - amount_paid, Decimal('0.00'))
            overpayment = max(amount_paid - grand_total, Decimal('0.00'))

            sale = form.save(commit=False)
            sale.payment_status = 'paid' if balance_due == 0 else ('unpaid' if amount_paid == 0 else 'partial')
            sale.balance_due = balance_due
            sale.save()

            for product, quantity in normalized_items:
                SaleItem.objects.create(
                    sale=sale,
                    product=product,
                    quantity=quantity,
                    unit_price=product.price,
                    unit_cost=product.cost,
                )
                if product.track_inventory:
                    product.stock = F('stock') - quantity
                    product.save(update_fields=['stock'])
                    StockMovement.objects.create(
                        product=product,
                        branch=sale.branch,
                        movement_type='sale',
                        quantity=-quantity,
                        reference=f'Sale #{sale.pk}',
                        created_by=request.user,
                    )

            customer = sale.customer
            if customer:
                customer.loyalty_points = F('loyalty_points') + total_profit_points
                update_fields = ['loyalty_points']
                if balance_due > 0:
                    customer.debt_balance = F('debt_balance') + balance_due
                    update_fields.append('debt_balance')
                elif overpayment > 0:
                    customer.credit_balance = F('credit_balance') + overpayment
                    update_fields.append('credit_balance')
                customer.save(update_fields=update_fields)

            Payment.objects.create(
                branch=sale.branch,
                sale=sale,
                method=sale.payment_method,
                amount=amount_paid,
                status='matched' if sale.payment_status == 'paid' else 'pending',
                transaction_reference=sale.reference,
            )
            ActivityLog.objects.create(profile=profile, branch=sale.branch, actor=str(sale.cashier or 'Cashier'), action=f'Recorded sale #{sale.pk} with {len(normalized_items)} item(s)')
            messages.success(request, f'Sale #{sale.pk} recorded and inventory updated.')
            return redirect(sale)
    else:
        form = SaleForm(profile=profile)

    context = build_pos_context(profile, form)
    return render(request, 'core/pos.html', context)


def build_pos_context(profile, form):
    return {
        'form': form,
        'products': Product.objects.select_related('category', 'branch').filter(branch__profile=profile, is_active=True),
        'product_options': [
            {
                'id': product.id,
                'name': product.name,
                'branch_id': product.branch_id,
                'stock': product.stock,
                'price': float(product.price),
                'track_inventory': product.track_inventory,
                'item_type': product.item_type,
            }
            for product in Product.objects.filter(branch__profile=profile, is_active=True).order_by('name')
        ],
        'cashier_options': [
            {
                'id': employee.id,
                'name': employee.name,
                'branch_id': employee.branch_id,
                'role': employee.get_role_display(),
            }
            for employee in Employee.objects.filter(branch__profile=profile, is_active=True).order_by('name')
        ],
        'today_sales': Sale.objects.filter(branch__profile=profile, created_at__date=timezone.localdate()).prefetch_related('items'),
    }


def sale_detail(request, pk):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_sales'):
        return permission_denied('manage_sales')
    sale = get_object_or_404(Sale.objects.filter(branch__profile=profile).select_related('branch', 'customer', 'cashier').prefetch_related('items__product'), pk=pk)
    receipt_message = build_receipt_message(profile, sale)
    whatsapp_url = ''
    if sale.customer and sale.customer.phone:
        whatsapp_url = build_whatsapp_url(sale.customer.phone, receipt_message)
    return render(request, 'core/sale_detail.html', {
        'sale': sale,
        'receipt_message': receipt_message,
        'whatsapp_url': whatsapp_url,
    })


def inventory(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_inventory'):
        return permission_denied('manage_inventory')
    query = request.GET.get('q', '')
    products = Product.objects.filter(branch__profile=profile).select_related('category', 'branch')
    if query:
        products = products.filter(name__icontains=query)
    return render(request, 'core/inventory.html', {'products': products, 'query': query})


def customers(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_customers'):
        return permission_denied('manage_customers')
    rows = []
    for customer in Customer.objects.filter(profile=profile).order_by('-loyalty_points', 'name'):
        rows.append({
            'customer': customer,
            'whatsapp_url': build_whatsapp_url(customer.phone, f'Hello {customer.name}, this is {profile.name}.') if customer.phone else '',
        })
    return render(request, 'core/customers.html', {'customer_rows': rows})


def employees(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_team'):
        return permission_denied('manage_team')
    return render(request, 'core/employees.html', {'employees': Employee.objects.filter(branch__profile=profile).select_related('branch')})


def hospitality(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_business'):
        return permission_denied('manage_business')
    units = HospitalityUnit.objects.filter(branch__profile=profile).select_related('branch')
    return render(request, 'core/hospitality.html', {'units': units})


def reports(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'view_reports'):
        return permission_denied('view_reports')
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="sms-sales-report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Receipt', 'Branch', 'Customer', 'Payment', 'Status', 'Total', 'Date'])
        for sale in Sale.objects.filter(branch__profile=profile).select_related('branch', 'customer').prefetch_related('items'):
            writer.writerow([
                sale.id,
                sale.branch.name,
                sale.customer.name if sale.customer else 'Walk-in',
                sale.get_payment_method_display(),
                sale.get_payment_status_display(),
                sale.total,
                sale.created_at.strftime('%Y-%m-%d %H:%M'),
            ])
        return response

    top_products = Product.objects.filter(branch__profile=profile).annotate(units_sold=Sum('sale_items__quantity')).order_by('-units_sold')[:8]
    sales_by_branch = Branch.objects.filter(profile=profile).annotate(sale_count=Count('sales')).order_by('name')
    expense_total = money(Expense.objects.filter(branch__profile=profile).aggregate(total=Sum('amount'))['total'])
    debt_total = money(Customer.objects.filter(profile=profile).aggregate(total=Sum('debt_balance'))['total'])
    context = {
        'sales_total': money(SaleItem.objects.filter(sale__branch__profile=profile).aggregate(total=Sum(F('quantity') * F('unit_price')))['total']),
        'profit_total': money(SaleItem.objects.filter(sale__branch__profile=profile).aggregate(total=Sum(F('quantity') * (F('unit_price') - F('unit_cost'))))['total']),
        'expense_total': expense_total,
        'net_profit': money(SaleItem.objects.filter(sale__branch__profile=profile).aggregate(total=Sum(F('quantity') * (F('unit_price') - F('unit_cost'))))['total']) - expense_total,
        'debt_total': debt_total,
        'top_products': top_products,
        'sales_by_branch': sales_by_branch,
        'payment_methods': Sale.objects.filter(branch__profile=profile).values('payment_method').annotate(total=Count('id')).order_by('-total'),
        'recent_expenses': Expense.objects.filter(branch__profile=profile).select_related('branch')[:8],
        'recent_purchases': Purchase.objects.filter(profile=profile).select_related('branch', 'supplier')[:8],
    }
    return render(request, 'core/reports.html', context)


def integrations(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_business'):
        return permission_denied('manage_business')
    ensure_default_connections(profile)
    if request.method == 'POST':
        form = BusinessProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            ActivityLog.objects.create(profile=profile, action='Updated business setup and integration settings')
            messages.success(request, 'Business setup saved.')
            return redirect('integrations')
    else:
        form = BusinessProfileForm(instance=profile)

    return render(request, 'core/integrations.html', {'form': form, 'profile': profile, 'connections': profile.connections.all()})


def connect_integration(request, connection_type):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_business'):
        return permission_denied('manage_business')
    ensure_default_connections(profile)
    connection = get_object_or_404(IntegrationConnection, profile=profile, connection_type=connection_type)
    if request.method == 'POST':
        form = IntegrationConnectionForm(request.POST, instance=connection)
        if form.is_valid():
            connection = form.save(commit=False)
            connection.status = 'connected'
            connection.save()
            ActivityLog.objects.create(profile=profile, action=f'Connected {connection.display_name}')
            messages.success(request, f'{connection.display_name} connected.')
            return redirect('integrations')
    else:
        form = IntegrationConnectionForm(instance=connection)
    return render(request, 'core/connect_integration.html', {'form': form, 'connection': connection})


def sync_integration(request, connection_type):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_business'):
        return permission_denied('manage_business')
    ensure_default_connections(profile)
    connection = get_object_or_404(IntegrationConnection, profile=profile, connection_type=connection_type)

    if connection_type == 'offline_excel':
        return redirect('import_excel')

    if connection_type not in ['google_sheets', 'excel_online']:
        messages.info(request, f'{connection.display_name} is connected for action links and manual payment workflows.')
        return redirect('connect_integration', connection_type=connection_type)

    branch = Branch.objects.filter(profile=profile, is_active=True).first()
    if not branch:
        messages.error(request, 'Add a branch before syncing spreadsheet data.')
        return redirect('connect_integration', connection_type=connection_type)

    try:
        rows = read_url_rows(connection.account_identifier)
        imported = import_product_rows(profile, branch, rows)
        connection.status = 'connected'
        connection.last_sync_at = timezone.now()
        connection.save(update_fields=['status', 'last_sync_at'])
        ActivityLog.objects.create(profile=profile, branch=branch, action=f'Synced {imported} products from {connection.display_name}')
        messages.success(request, f'Synced {imported} products from {connection.display_name}.')
    except Exception:
        messages.error(request, 'Could not sync that spreadsheet. Use a public CSV/XLSX link or upload the file from Excel Import.')
    return redirect('connect_integration', connection_type=connection_type)


def import_excel(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_inventory'):
        return permission_denied('manage_inventory')
    imported = 0
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES, profile=profile)
        if form.is_valid():
            branch = form.cleaned_data['branch']
            upload = form.cleaned_data['file']
            rows = read_upload_rows(upload)
            imported = import_product_rows(profile, branch, rows)
            ActivityLog.objects.create(profile=profile, branch=branch, action=f'Imported {imported} products from spreadsheet')
            messages.success(request, f'Imported or updated {imported} products.')
            return redirect('inventory')
    else:
        form = ExcelImportForm(profile=profile)
    return render(request, 'core/import_excel.html', {'form': form, 'imported': imported})


def export_excel(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_inventory'):
        return permission_denied('manage_inventory')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Products'
    sheet.append(['Product Name', 'SKU', 'Branch', 'Category', 'Supplier', 'Qty', 'Cost Price', 'Selling Price', 'Barcode'])
    for product in Product.objects.filter(branch__profile=profile).select_related('branch', 'category', 'supplier'):
        sheet.append([
            product.name,
            product.sku,
            product.branch.name,
            product.category.name,
            product.supplier.name if product.supplier else '',
            product.stock,
            product.cost,
            product.price,
            product.barcode,
        ])
    output = io.BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sms-products-offline.xlsx"'
    return response


def payments(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_payments'):
        return permission_denied('manage_payments')
    grouped = Payment.objects.filter(branch__profile=profile).values('method', 'status').annotate(total=Sum('amount'), count=Count('id')).order_by('method')
    return render(request, 'core/payments.html', {'payments': Payment.objects.filter(branch__profile=profile).select_related('branch', 'sale'), 'grouped': grouped})


def team(request):
    profile = get_profile_for_request(request)
    ensure_default_roles_and_permissions()
    if not user_has_business_permission(request, 'manage_team'):
        return HttpResponse('You do not have permission to manage the team.', status=403)

    if request.method == 'POST':
        form = TeamMemberForm(request.POST, profile=profile)
        if form.is_valid():
            email = form.cleaned_data['email'].strip().lower()
            user = User.objects.filter(email__iexact=email).first()
            access, created = UserBusinessAccess.objects.update_or_create(
                email=email,
                business=profile,
                defaults={
                    'user': user,
                    'role': form.cleaned_data['role'],
                    'is_active': form.cleaned_data['is_active'],
                },
            )
            ActivityLog.objects.create(profile=profile, actor=request.user.get_username(), action=f'Updated team access for {email}')
            if created or not user:
                registration_url = f"{settings.SITE_URL}/auth/register/"
                send_mail(
                    subject=f'Invitation to join {profile.name} on SMS Suite',
                    message=(
                        f'You have been invited to join {profile.name} on SMS Suite.\n\n'
                        f'Role: {access.role.name}\n\n'
                        f'Click the link below to create your account:\n{registration_url}\n\n'
                        f'After registering, you will automatically have access to the business.'
                    ),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    fail_silently=True,
                )
            messages.success(request, 'Team member access saved.')
            return redirect('team')
    else:
        form = TeamMemberForm(profile=profile, initial={'is_active': True})

    accesses = UserBusinessAccess.objects.filter(business=profile).select_related('user', 'role')
    return render(request, 'core/team.html', {'form': form, 'accesses': accesses, 'profile': profile})


def crud_list(request, resource):
    profile = get_profile_for_request(request)
    permission_code = resource_permission(resource)
    if not user_has_business_permission(request, permission_code):
        return permission_denied(permission_code)
    config = get_crud_config(resource)
    items = scoped_queryset(config['model'], profile)
    return render(request, 'core/crud_list.html', {'resource': resource, 'config': config, 'items': items})


def crud_create(request, resource):
    profile = get_profile_for_request(request)
    permission_code = resource_permission(resource)
    if not user_has_business_permission(request, permission_code):
        return permission_denied(permission_code)
    config = get_crud_config(resource)
    form_class = CRUD_FORMS[resource]
    if request.method == 'POST':
        form = form_class(request.POST, profile=profile)
        if form.is_valid():
            item = form.save(commit=False)
            attach_profile(item, profile)
            item.save()
            form.save_m2m()
            ActivityLog.objects.create(profile=profile, action=f'Created {config["title"]}: {item}')
            messages.success(request, f'{config["title"]} item created.')
            return redirect('crud_list', resource=resource)
    else:
        form = form_class(profile=profile)
    return render(request, 'core/crud_form.html', {'resource': resource, 'config': config, 'form': form, 'mode': 'Add'})


def crud_update(request, resource, pk):
    profile = get_profile_for_request(request)
    permission_code = resource_permission(resource)
    if not user_has_business_permission(request, permission_code):
        return permission_denied(permission_code)
    config = get_crud_config(resource)
    item = get_object_or_404(scoped_queryset(config['model'], profile), pk=pk)
    form_class = CRUD_FORMS[resource]
    if request.method == 'POST':
        form = form_class(request.POST, instance=item, profile=profile)
        if form.is_valid():
            item = form.save()
            ActivityLog.objects.create(profile=profile, action=f'Updated {config["title"]}: {item}')
            messages.success(request, f'{config["title"]} item updated.')
            return redirect('crud_list', resource=resource)
    else:
        form = form_class(instance=item, profile=profile)
    return render(request, 'core/crud_form.html', {'resource': resource, 'config': config, 'form': form, 'mode': 'Edit', 'item': item})


def crud_delete(request, resource, pk):
    profile = get_profile_for_request(request)
    permission_code = resource_permission(resource)
    if not user_has_business_permission(request, permission_code):
        return permission_denied(permission_code)
    config = get_crud_config(resource)
    item = get_object_or_404(scoped_queryset(config['model'], profile), pk=pk)
    if request.method == 'POST':
        item.delete()
        ActivityLog.objects.create(profile=profile, action=f'Deleted {config["title"]}: {item}')
        messages.success(request, f'{config["title"]} item deleted.')
        return redirect('crud_list', resource=resource)
    return render(request, 'core/crud_confirm_delete.html', {'resource': resource, 'config': config, 'item': item})


def get_crud_config(resource):
    if resource not in CRUD_CONFIG:
        raise Http404
    return CRUD_CONFIG[resource]


def scoped_queryset(model, profile):
    if model in [Branch, Category, Supplier, Customer]:
        return model.objects.filter(profile=profile)
    if model is Product:
        return Product.objects.filter(branch__profile=profile)
    if model is Employee:
        return Employee.objects.filter(branch__profile=profile)
    if model is HospitalityUnit:
        return HospitalityUnit.objects.filter(branch__profile=profile)
    if model in [Payment, Expense]:
        return model.objects.filter(branch__profile=profile)
    if model is StockTransfer:
        return StockTransfer.objects.filter(from_branch__profile=profile)
    return model.objects.none()


def attach_profile(item, profile):
    if hasattr(item, 'profile_id'):
        item.profile = profile


def ensure_default_connections(profile):
    defaults = {
        'google_sheets': ('Google Sheets', 'Paste a Google Sheet link or account email'),
        'excel_online': ('Microsoft Excel Online', 'Paste OneDrive workbook link'),
        'whatsapp': ('WhatsApp Business', profile.whatsapp_number or 'Add WhatsApp number'),
        'offline_excel': ('Offline Excel', 'Import/export local Excel files'),
        'mtn_momo': ('MTN MoMo', profile.momo_number or 'Add MoMo number'),
        'airtel_money': ('Airtel Money', profile.airtel_money_number or 'Add Airtel number'),
    }
    for key, (name, identifier) in defaults.items():
        IntegrationConnection.objects.get_or_create(
            profile=profile,
            connection_type=key,
            defaults={'display_name': name, 'account_identifier': identifier, 'status': 'needs_setup'},
        )


def expenses(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_payments'):
        return permission_denied('manage_payments')
    if request.method == 'POST':
        form = ExpenseForm(request.POST, profile=profile)
        if form.is_valid():
            expense = form.save()
            ActivityLog.objects.create(profile=profile, branch=expense.branch, action=f'Recorded expense: {expense.title}')
            messages.success(request, 'Expense recorded.')
            return redirect('expenses')
    else:
        form = ExpenseForm(profile=profile)
    return render(request, 'core/expenses.html', {'form': form, 'expenses': Expense.objects.filter(branch__profile=profile).select_related('branch')})


def transfers(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_inventory'):
        return permission_denied('manage_inventory')
    if request.method == 'POST':
        form = StockTransferForm(request.POST, profile=profile)
        if form.is_valid():
            quantity = form.cleaned_data['quantity']
            with transaction.atomic():
                product = Product.objects.select_for_update().get(pk=form.cleaned_data['product'].pk)
                if product.stock < quantity:
                    form.add_error('quantity', f'Only {product.stock} units are available.')
                    return render(request, 'core/transfers.html', {'form': form, 'transfers': StockTransfer.objects.filter(from_branch__profile=profile).select_related('from_branch', 'to_branch')})
                product.stock = F('stock') - quantity
                product.save(update_fields=['stock'])
                destination, _ = Product.objects.get_or_create(
                    branch=form.cleaned_data['to_branch'],
                    sku=product.sku,
                    defaults={
                        'name': product.name,
                        'item_type': product.item_type,
                        'category': product.category,
                        'supplier': product.supplier,
                        'barcode': product.barcode,
                        'image_url': product.image_url,
                        'price': product.price,
                        'cost': product.cost,
                        'stock': 0,
                        'reorder_level': product.reorder_level,
                        'track_inventory': product.track_inventory,
                        'batch_number': product.batch_number,
                        'expiry_date': product.expiry_date,
                        'is_active': product.is_active,
                    },
                )
                destination.stock = F('stock') + quantity
                destination.save(update_fields=['stock'])
                transfer = form.save(commit=False)
                transfer.product = product
                transfer.product_name = product.name
                transfer.save()
                StockMovement.objects.create(product=product, branch=product.branch, movement_type='transfer_out', quantity=-quantity, reference=f'Transfer #{transfer.pk}', notes=transfer.notes, created_by=request.user)
                StockMovement.objects.create(product=destination, branch=destination.branch, movement_type='transfer_in', quantity=quantity, reference=f'Transfer #{transfer.pk}', notes=transfer.notes, created_by=request.user)
            ActivityLog.objects.create(profile=profile, branch=transfer.from_branch, action=f'Stock transfer requested: {transfer.product_name}')
            messages.success(request, 'Stock transfer recorded.')
            return redirect('transfers')
    else:
        form = StockTransferForm(profile=profile)
    return render(request, 'core/transfers.html', {'form': form, 'transfers': StockTransfer.objects.filter(from_branch__profile=profile).select_related('from_branch', 'to_branch', 'product')})


def stock_movements(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_inventory'):
        return permission_denied('manage_inventory')
    if request.method == 'POST':
        form = StockMovementForm(request.POST, profile=profile)
        if form.is_valid():
            quantity = form.cleaned_data['quantity']
            movement_type = form.cleaned_data['movement_type']
            signed_quantity = quantity if movement_type == 'stock_in' else -quantity
            with transaction.atomic():
                product = Product.objects.select_for_update().get(pk=form.cleaned_data['product'].pk)
                if movement_type in ['stock_out', 'adjustment'] and product.stock < quantity:
                    form.add_error('quantity', f'Only {product.stock} units are available.')
                else:
                    product.stock = F('stock') + signed_quantity
                    product.save(update_fields=['stock'])
                    StockMovement.objects.create(
                        product=product,
                        branch=product.branch,
                        movement_type=movement_type,
                        quantity=signed_quantity,
                        notes=form.cleaned_data['notes'],
                        created_by=request.user,
                    )
                    ActivityLog.objects.create(profile=profile, branch=product.branch, actor=request.user.get_username(), action=f'Recorded {movement_type.replace("_", " ")} for {product.name}')
                    messages.success(request, 'Stock movement recorded.')
                    return redirect('stock_movements')
    else:
        form = StockMovementForm(profile=profile)
    movements = StockMovement.objects.filter(branch__profile=profile).select_related('product', 'branch', 'created_by')[:80]
    return render(request, 'core/stock_movements.html', {'form': form, 'movements': movements})



def refresh_business_alerts(profile):
    today = timezone.localdate()
    for product in Product.objects.filter(branch__profile=profile, is_active=True, track_inventory=True):
        if product.stock <= product.reorder_level:
            Notification.objects.get_or_create(
                profile=profile,
                branch=product.branch,
                title=f'Low stock: {product.name}',
                defaults={
                    'message': f'{product.name} has {product.stock} left at {product.branch.name}. Reorder level is {product.reorder_level}.',
                    'level': 'warning',
                    'action_url': reverse('crud_list', args=['products']),
                },
            )
        if product.expiry_date and product.expiry_date <= today + timedelta(days=30):
            Notification.objects.get_or_create(
                profile=profile,
                branch=product.branch,
                title=f'Expiry alert: {product.name}',
                defaults={
                    'message': f'{product.name} batch {product.batch_number or "N/A"} expires on {product.expiry_date}.',
                    'level': 'danger',
                    'action_url': reverse('stock_movements'),
                },
            )
    for customer in Customer.objects.filter(profile=profile, debt_balance__gt=0):
        Notification.objects.get_or_create(
            profile=profile,
            title=f'Debt reminder: {customer.name}',
            defaults={
                'message': f'{customer.name} owes RWF {customer.debt_balance}.',
                'level': 'info',
                'action_url': reverse('customers'),
            },
        )
    if profile.trial_ends_at and profile.trial_days_left <= 14:
        Notification.objects.get_or_create(
            profile=profile,
            title='Trial ending soon',
            defaults={
                'message': f'Your free trial has {profile.trial_days_left} days left. Prepare billing before real customer launch.',
                'level': 'warning',
                'action_url': reverse('integrations'),
            },
        )


def notifications(request):
    profile = get_profile_for_request(request)
    refresh_business_alerts(profile)
    if request.method == 'POST':
        Notification.objects.filter(profile=profile).update(is_read=True)
        messages.success(request, 'All notifications marked as read.')
        return redirect('notifications')
    return render(request, 'core/notifications.html', {
        'notifications': Notification.objects.filter(profile=profile).select_related('branch')[:100],
    })


@transaction.atomic
def purchases(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_inventory'):
        return permission_denied('manage_inventory')
    if request.method == 'POST':
        form = PurchaseForm(request.POST, profile=profile)
        if form.is_valid():
            product = Product.objects.select_for_update().get(pk=form.cleaned_data['product'].pk)
            qty = form.cleaned_data['quantity']
            unit_cost = form.cleaned_data['unit_cost']
            purchase = form.save(commit=False)
            purchase.profile = profile
            purchase.created_by = request.user
            purchase.total_cost = qty * unit_cost
            purchase.status = 'received'
            purchase.save()
            PurchaseItem.objects.create(purchase=purchase, product=product, quantity=qty, unit_cost=unit_cost)
            product.stock = F('stock') + qty
            product.cost = unit_cost
            product.save(update_fields=['stock', 'cost'])
            StockMovement.objects.create(product=product, branch=product.branch, movement_type='stock_in', quantity=qty, reference=f'Purchase #{purchase.pk}', notes=purchase.notes, created_by=request.user)
            ActivityLog.objects.create(profile=profile, branch=product.branch, actor=request.user.get_username(), action=f'Received purchase #{purchase.pk}: {product.name} x{qty}')
            messages.success(request, 'Purchase received and stock updated.')
            return redirect('purchases')
    else:
        form = PurchaseForm(profile=profile)
    return render(request, 'core/purchases.html', {
        'form': form,
        'purchases': Purchase.objects.filter(profile=profile).select_related('branch', 'supplier', 'created_by').prefetch_related('items__product')[:80],
    })


def daily_close(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_payments'):
        return permission_denied('manage_payments')
    if request.method == 'POST':
        form = CashCloseForm(request.POST, profile=profile)
        if form.is_valid():
            branch = form.cleaned_data['branch']
            opening_cash = form.cleaned_data['opening_cash']
            closing_cash = form.cleaned_data['closing_cash']
            cash_sales = money(Payment.objects.filter(branch=branch, method='cash', recorded_at__date=timezone.localdate()).aggregate(total=Sum('amount'))['total'])
            expected = opening_cash + cash_sales
            session = CashRegisterSession.objects.create(
                branch=branch,
                cashier=form.cleaned_data.get('cashier'),
                opened_by=request.user,
                opening_cash=opening_cash,
                closing_cash=closing_cash,
                expected_cash=expected,
                difference=closing_cash - expected,
                status='closed',
                closed_at=timezone.now(),
                notes=form.cleaned_data['notes'],
            )
            ActivityLog.objects.create(profile=profile, branch=branch, actor=request.user.get_username(), action=f'Closed cash drawer #{session.pk}')
            messages.success(request, 'Daily cash closing saved.')
            return redirect('daily_close')
    else:
        form = CashCloseForm(profile=profile)
    return render(request, 'core/daily_close.html', {
        'form': form,
        'sessions': CashRegisterSession.objects.filter(branch__profile=profile).select_related('branch', 'cashier')[:50],
    })


def backup_center(request):
    profile = get_profile_for_request(request)
    if not user_has_business_permission(request, 'manage_business'):
        return permission_denied('manage_business')
    logs = BackupLog.objects.filter(Q(profile=profile) | Q(profile__isnull=True))[:30]
    return render(request, 'core/backup_center.html', {'logs': logs, 'profile': profile})


def read_upload_rows(upload):
    name = upload.name.lower()
    if name.endswith('.csv'):
        text = upload.read().decode('utf-8-sig')
        return list(csv.DictReader(io.StringIO(text)))

    workbook = load_workbook(upload, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or '').strip() for value in rows[0]]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in rows[1:]
    ]


def read_url_rows(url):
    if not url:
        raise ValueError('No spreadsheet URL configured.')
    normalized_url = normalize_spreadsheet_url(url)
    response = requests.get(normalized_url, timeout=20)
    response.raise_for_status()
    content_type = response.headers.get('content-type', '')
    if normalized_url.lower().endswith('.xlsx') or 'spreadsheetml' in content_type:
        workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or '').strip() for value in rows[0]]
        return [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in rows[1:]
        ]
    text = response.content.decode('utf-8-sig')
    return list(csv.DictReader(io.StringIO(text)))


def normalize_spreadsheet_url(url):
    if 'docs.google.com/spreadsheets' in url and '/edit' in url:
        sheet_id = url.split('/d/')[1].split('/')[0]
        return f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv'
    if '1drv.ms' in url or 'onedrive.live.com' in url:
        return url
    return url


def import_product_rows(profile, branch, rows):
    imported = 0
    for row in rows:
        name = pick_value(row, ['product name', 'item name', 'name', 'product'])
        if not name:
            continue
        category_name = pick_value(row, ['category'], 'Imported')
        category, _ = Category.objects.get_or_create(profile=profile, name=category_name or 'Imported')
        supplier_name = pick_value(row, ['supplier', 'vendor'])
        supplier = None
        if supplier_name:
            supplier, _ = Supplier.objects.get_or_create(profile=profile, name=supplier_name)
        sku = pick_value(row, ['sku', 'code', 'barcode']) or f'IMP-{branch.id}-{imported + 1:04d}'
        Product.objects.update_or_create(
            branch=branch,
            sku=sku,
            defaults={
                'name': name,
                'branch': branch,
                'category': category,
                'supplier': supplier,
                'barcode': pick_value(row, ['barcode'], ''),
                'stock': int(decimal_value(pick_value(row, ['qty', 'stock', 'quantity'], 0))),
                'cost': decimal_value(pick_value(row, ['cost price', 'buying price', 'cost'], 0)),
                'price': decimal_value(pick_value(row, ['selling price', 'price', 'sale price'], 0)),
            },
        )
        imported += 1
    return imported


def pick_value(row, names, default=''):
    normalized = {str(key).strip().lower(): value for key, value in row.items()}
    for name in names:
        value = normalized.get(name)
        if value not in (None, ''):
            return str(value).strip()
    return default


def decimal_value(value):
    try:
        return Decimal(str(value or 0).replace(',', ''))
    except Exception:
        return Decimal('0.00')


def build_whatsapp_url(phone, message):
    clean_phone = ''.join(char for char in phone if char.isdigit())
    if clean_phone.startswith('0'):
        clean_phone = '250' + clean_phone[1:]
    return f'https://wa.me/{clean_phone}?{urlencode({"text": message})}'


def build_receipt_message(profile, sale):
    lines = [
        f'{profile.name} receipt #{sale.id}',
        f'Branch: {sale.branch.name}',
        f'Subtotal: R{sale.subtotal:.2f}',
        f'Discount: R{sale.discount_amount:.2f}',
        f'Total: R{sale.grand_total:.2f}',
        f'Paid: R{sale.amount_paid:.2f}',
        f'Balance: R{sale.balance_due:.2f}',
        f'Payment: {sale.get_payment_method_display()} ({sale.get_payment_status_display()})',
    ]
    for item in sale.items.all():
        lines.append(f'- {item.product.name} x{item.quantity}: R{item.line_total:.2f}')
    lines.append('Thank you for your business.')
    return '\n'.join(lines)

# Create your views here.
